#!/usr/bin/env python3
"""
BOPS PPT 週次統計 Excel 生成スクリプト v3 (Dashboard + 明細 デュアル Sheet)

使い方:
    python3 generate_xlsx.py <input.tsv> <output.xlsx> [<start_date> <end_date>]

入力 TSV 形式は 2 種類対応:
  v1 (14 列): id topic username companyName themePresetId strategyName status
              slideCount actualSlides createdAt startTime endTime durationSec logsCount
  v2 (17 列): 上記 14 列 + originalDurationSec breakCount maxGapSec

出力 Excel:
  Sheet 1 「ダッシュボード」: タイトル / KPI 4枚 / グラフ 4枚 / 完成度ブロック
  Sheet 2 「明細データ」: 全件詳細表 + 合計・平均行

統計用スライド数 (J列):
    status が "" or "created" && logsCount ≤ 1  →  0
    それ以外                                      →  実際スライド数 I

GAP 閾値: 1 時間（3600 秒）以上の log 間隔は「ユーザー中断」と判定（v2 のみ）
"""

import csv
import sys
from collections import Counter
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, BarChart3D, LineChart, ScatterChart, PieChart, Reference, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import SeriesLabel, StrRef
from openpyxl.chart.marker import Marker
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.fill import ColorChoice

HEADER_V1 = [
    'ID', 'トピック', 'ユーザー', '会社', 'テーマ', '戦略',
    '状態', 'スライド数（一覧）', '実際スライド数', '統計用スライド数',
    '作成日時', '開始時刻', '終了時刻',
    '所要時間（秒）', '所要時間（mm:ss）', 'ログ件数',
    '平均所用時間／１頁（秒）', '未做成原因', '改善方案'
]
HEADER_V2 = HEADER_V1 + ['元の所要時間（秒）', '中断回数', '最大中断時間（秒）']
COL_WIDTHS_V1 = [7, 42, 11, 23, 7, 32, 21, 9, 13, 13, 18, 18, 18, 13, 13, 9, 15, 31, 25]
COL_WIDTHS_V2 = COL_WIDTHS_V1 + [14, 10, 14]

# 配色（プロフェッショナルな見た目に）
COLOR_TITLE_BG = '1F4E78'      # 濃紺
COLOR_TITLE_FG = 'FFFFFF'
COLOR_KPI_BG = 'D9E1F2'        # 薄水色
COLOR_KPI_LABEL = '305496'
COLOR_HEADER_BG = '305496'
COLOR_BREAK_BG = 'FFE699'      # 薄黄（中断あり）
COLOR_EXCLUDED_BG = 'F2F2F2'   # 薄灰（除外）
COLOR_TOTAL_BG = 'FFF2CC'      # 薄黄（合計行）
COLOR_AVG_BG = 'FCE4D6'        # 薄橙（平均行）


def classify(status: str, logs_count: int) -> tuple[int, str]:
    s = (status or '').strip().lower()
    if logs_count <= 1 and s in ('', 'created'):
        if s == 'created':
            return 0, 'ユーザーが処理を中断'
        return 0, '未生成（要調査）'
    return 1, ''


def fmt_mmss(sec: int) -> str:
    return f'{sec // 60:02d}:{sec % 60:02d}'


def fmt_minsec(sec: int) -> str:
    if sec < 60:
        return f'{sec}秒'
    m, s = sec // 60, sec % 60
    if m < 60:
        return f'{m}分{s:02d}秒'
    h, m = m // 60, m % 60
    return f'{h}時間{m}分'


def aggregate(records):
    """期間集計メトリクスを返す。当期・前期で共通利用し、前期比（環比）を出す。

    返り値の主要キー:
      total          総件数
      generated      生成成功件数（J>0）
      excluded       未生成件数（J=0）
      completion     完成率 0..1
      avg_duration   1件あたり平均純所要（秒、生成成功のみ）
      avg_per_slide  1ページあたり平均（秒、生成成功のみ）
      total_slides   生成成功分の総ページ数
      companies      Counter(会社 -> 件数)  ← 使用意愿度シグナルの元データ
      users          Counter(ユーザー -> 件数)
    """
    total = len(records)
    gen = [r for r in records if classify(r['status'], r['logsCount'])[0] == 1]
    generated = len(gen)
    excluded = total - generated
    total_duration = sum(r['durationSec'] for r in gen)
    total_slides = sum(r['actualSlides'] for r in gen)
    return {
        'total': total,
        'generated': generated,
        'excluded': excluded,
        'completion': (generated / total) if total else 0,
        'avg_duration': (total_duration // generated) if generated else 0,
        'avg_per_slide': (total_duration / total_slides) if total_slides else 0,
        'total_slides': total_slides,
        'companies': Counter(r['company'] for r in records),
        'users': Counter(r['username'] for r in records),
    }


def fmt_delta(cur, prev, unit='', pct=False):
    """前期比の差分を ▲/▼ 付き文字列で返す。prev が無ければ '—'。"""
    if prev is None:
        return '—'
    diff = cur - prev
    if pct:
        cur_s, diff_s = f'{cur*100:.1f}', f'{abs(diff)*100:.1f}'
    else:
        cur_s, diff_s = f'{cur:g}', f'{abs(diff):g}'
    if diff > 0:
        return f'▲ +{diff_s}{unit}'
    if diff < 0:
        return f'▼ -{diff_s}{unit}'
    return f'± 0{unit}'


def parse_tsv(input_tsv: str):
    """Load TSV → list of dicts. Returns (records, is_v2)."""
    with open(input_tsv, encoding='utf-8') as f:
        rows = list(csv.reader(f, delimiter='\t'))
    if not rows:
        raise ValueError('Empty TSV')

    cols = len(rows[0])
    if cols == 14:
        is_v2 = False
    elif cols == 17:
        is_v2 = True
    else:
        raise ValueError(f'Unexpected TSV column count: {cols}')

    records = []
    for r in rows[1:]:
        if is_v2:
            (id_, topic, user, company, theme, strategy, status,
             slide_cnt, actual_slides, created, start, end, dur_sec, log_cnt,
             orig_dur, break_cnt, max_gap) = r
        else:
            (id_, topic, user, company, theme, strategy, status,
             slide_cnt, actual_slides, created, start, end, dur_sec, log_cnt) = r
            orig_dur = dur_sec
            break_cnt = '0'
            max_gap = '0'
        records.append({
            'id': int(id_), 'topic': topic, 'username': user, 'company': company,
            'theme': theme, 'strategy': strategy, 'status': status,
            'slideCount': int(slide_cnt), 'actualSlides': int(actual_slides),
            'createdAt': created, 'startTime': start, 'endTime': end,
            'durationSec': int(dur_sec), 'logsCount': int(log_cnt),
            'originalDurationSec': int(orig_dur),
            'breakCount': int(break_cnt), 'maxGapSec': int(max_gap),
        })
    return records, is_v2


def build_detail_sheet(ws, records, is_v2):
    """明細データ Sheet を作る。"""
    header = HEADER_V2 if is_v2 else HEADER_V1
    widths = COL_WIDTHS_V2 if is_v2 else COL_WIDTHS_V1
    ws.append(header)

    thin = Side(border_style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', start_color=COLOR_HEADER_BG)
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    excluded_fill = PatternFill('solid', start_color=COLOR_EXCLUDED_BG)
    break_fill = PatternFill('solid', start_color=COLOR_BREAK_BG)

    for c in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
    ws.row_dimensions[1].height = 32

    excluded_count = 0
    break_count_total = 0

    for idx, rec in enumerate(records, start=2):
        j_mult, default_reason = classify(rec['status'], rec['logsCount'])
        j_value = rec['actualSlides'] * j_mult
        if j_value == 0:
            excluded_count += 1
        if rec['breakCount'] > 0:
            break_count_total += 1
            if default_reason:
                default_reason += f' / 中断 {rec["breakCount"]} 回（元 {rec["originalDurationSec"]}s → 純 {rec["durationSec"]}s）'
            else:
                default_reason = f'中断 {rec["breakCount"]} 回（元 {rec["originalDurationSec"]}s → 純 {rec["durationSec"]}s、最長 {fmt_mmss(rec["maxGapSec"])}）'

        avg_formula = f'=N{idx}/I{idx}' if rec['actualSlides'] > 0 else 0
        row_values = [
            rec['id'], rec['topic'], rec['username'], rec['company'], rec['theme'],
            rec['strategy'], rec['status'],
            rec['slideCount'], rec['actualSlides'], j_value,
            rec['createdAt'], rec['startTime'], rec['endTime'],
            rec['durationSec'], fmt_mmss(rec['durationSec']), rec['logsCount'],
            avg_formula, default_reason, ''
        ]
        if is_v2:
            row_values += [rec['originalDurationSec'], rec['breakCount'], rec['maxGapSec']]
        ws.append(row_values)

        for c in range(1, len(header) + 1):
            cell = ws.cell(row=idx, column=c)
            cell.border = border
            cell.alignment = Alignment(vertical='center', wrap_text=(c in (2, 18, 19)))
        if j_value == 0:
            for c in (10, 14, 15, 17):
                ws.cell(row=idx, column=c).fill = excluded_fill
        elif rec['breakCount'] > 0:
            cols_to_highlight = (14, 15, 17, 20, 21, 22) if is_v2 else (14, 15, 17)
            for c in cols_to_highlight:
                ws.cell(row=idx, column=c).fill = break_fill

    data_end = ws.max_row
    total_rows = data_end - 1

    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'

    # 合計行
    total_row = data_end + 1
    bold = Font(bold=True)
    total_fill = PatternFill('solid', start_color=COLOR_TOTAL_BG)
    avg_fill = PatternFill('solid', start_color=COLOR_AVG_BG)

    ws.cell(row=total_row, column=1, value='合計').font = bold
    ws.cell(row=total_row, column=8, value=f'=SUM(H2:H{data_end})').font = bold
    ws.cell(row=total_row, column=9, value=f'=SUM(I2:I{data_end})').font = bold
    ws.cell(row=total_row, column=10, value=f'=SUM(J2:J{data_end})').font = bold
    ws.cell(row=total_row, column=14, value=f'=SUM(N2:N{data_end})').font = bold
    ws.cell(row=total_row, column=17, value=f'=IF(J{total_row}=0,0,N{total_row}/J{total_row})').font = bold
    ws.cell(row=total_row, column=17).number_format = '0.0'
    if is_v2:
        ws.cell(row=total_row, column=20, value=f'=SUM(T2:T{data_end})').font = bold
        ws.cell(row=total_row, column=21, value=f'=SUM(U2:U{data_end})').font = bold
    for c in range(1, len(header) + 1):
        ws.cell(row=total_row, column=c).fill = total_fill
        ws.cell(row=total_row, column=c).border = border

    avg_row = total_row + 1
    bold_red = Font(bold=True, color='C00000')
    ws.cell(row=avg_row, column=1, value='平均').font = bold_red
    ws.cell(row=avg_row, column=8, value=f'=AVERAGE(H2:H{data_end})').font = bold_red
    ws.cell(row=avg_row, column=9, value=f'=AVERAGE(I2:I{data_end})').font = bold_red
    ws.cell(row=avg_row, column=10, value=f'=AVERAGE(J2:J{data_end})').font = bold_red
    ws.cell(row=avg_row, column=14, value=f'=AVERAGE(N2:N{data_end})').font = bold_red
    ws.cell(row=avg_row, column=14).number_format = '0.0'
    for c in range(1, len(header) + 1):
        ws.cell(row=avg_row, column=c).fill = avg_fill
        ws.cell(row=avg_row, column=c).border = border

    return {
        'total_rows': total_rows,
        'excluded_count': excluded_count,
        'break_count_total': break_count_total,
        'data_end_row': data_end,
        'total_row': total_row,
        'avg_row': avg_row,
    }


def build_dashboard(ws, records, is_v2, start_date, end_date, detail_stats,
                    prev_agg=None, mode='weekly', prev_label=''):
    """ダッシュボード Sheet を作る。

    prev_agg: aggregate() の前期メトリクス。渡されると前期比（環比）行と
              会社別トレンドに増減列が出る（使用意愿度シグナル）。
    mode:     'daily' / 'weekly' — タイトルと粒度ラベルに反映。
    """
    # ----- 列幅をセット（A〜L まで 14ぐらい） -----
    for col_letter in 'ABCDEFGHIJKL':
        ws.column_dimensions[col_letter].width = 14

    bold = Font(bold=True)
    title_font = Font(bold=True, color=COLOR_TITLE_FG, size=20)
    title_fill = PatternFill('solid', start_color=COLOR_TITLE_BG)
    subtitle_font = Font(italic=True, color='FFFFFF', size=11)
    section_font = Font(bold=True, color=COLOR_TITLE_BG, size=14)
    kpi_label_font = Font(bold=True, color=COLOR_KPI_LABEL, size=10)
    kpi_value_font = Font(bold=True, size=22, color='1F4E78')
    kpi_unit_font = Font(color='808080', size=9)
    kpi_fill = PatternFill('solid', start_color=COLOR_KPI_BG)
    thin = Side(border_style='thin', color='B4C7E7')
    thick = Side(border_style='medium', color='1F4E78')
    kpi_border = Border(left=thick, right=thick, top=thick, bottom=thick)
    center = Alignment(horizontal='center', vertical='center')

    # ----- 集計計算 -----
    total_rows = detail_stats['total_rows']
    excluded_count = detail_stats['excluded_count']
    break_count_total = detail_stats['break_count_total']
    generated = total_rows - excluded_count
    completion_rate = generated / total_rows if total_rows else 0

    # 平均所要時間（J>0 のみ対象）
    gen_records = [r for r in records if classify(r['status'], r['logsCount'])[0] == 1]
    total_duration = sum(r['durationSec'] for r in gen_records)
    avg_duration = total_duration // len(gen_records) if gen_records else 0
    total_slides = sum(r['actualSlides'] for r in gen_records)
    avg_per_slide = total_duration / total_slides if total_slides else 0
    break_rate = break_count_total / total_rows if total_rows else 0

    # ----- タイトル -----
    mode_label = '日次' if mode == 'daily' else '週次'
    ws.row_dimensions[1].height = 38
    ws.cell(row=1, column=1, value=f'BOPS PPT {mode_label}統計 ダッシュボード').font = title_font
    ws.cell(row=1, column=1).fill = title_fill
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)

    period_str = f'対象期間: {start_date} ~ {end_date}' if start_date else ''
    ws.cell(row=2, column=1, value=period_str).font = subtitle_font
    ws.cell(row=2, column=1).fill = title_fill
    ws.cell(row=2, column=1).alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=12)
    ws.row_dimensions[2].height = 18

    # ----- KPI 4 枚（row 4-6 を使う）。前期比（環比）をヒントに併記 -----
    pa = prev_agg
    d_total = fmt_delta(total_rows, pa['total']) if pa else None
    d_comp = fmt_delta(completion_rate, pa['completion'], pct=True) if pa else None
    d_dur = fmt_delta(avg_duration, pa['avg_duration'], unit='s') if pa else None
    d_slide = fmt_delta(round(avg_per_slide), round(pa['avg_per_slide']), unit='s') if pa else None
    pfx = f'前期比 ' if pa else ''
    kpis = [
        ('総生成件数', f'{total_rows}', '本', (f'{pfx}{d_total}' if pa else None)),
        ('完成率', f'{completion_rate*100:.1f}', '%',
         (f'生成成功 {generated}/{total_rows}　{d_comp}' if pa else f'生成成功 {generated} / {total_rows}')),
        ('1件あたり平均純所要', fmt_minsec(avg_duration), '',
         (f'対象 {len(gen_records)} 件　{d_dur}' if pa else f'集計対象 {len(gen_records)} 件')),
        ('1ページあたり平均', f'{avg_per_slide:.0f}', '秒',
         (f'総ページ {total_slides}　{d_slide}' if pa else f'総ページ {total_slides}')),
    ]
    kpi_row_label = 4
    kpi_row_value = 5
    kpi_row_unit = 6
    ws.row_dimensions[kpi_row_label].height = 18
    ws.row_dimensions[kpi_row_value].height = 36
    ws.row_dimensions[kpi_row_unit].height = 16

    for i, (label, value, unit, hint) in enumerate(kpis):
        c_start = 1 + i * 3  # A,D,G,J
        c_end = c_start + 2

        # ラベル行
        ws.cell(row=kpi_row_label, column=c_start, value=label).font = kpi_label_font
        ws.cell(row=kpi_row_label, column=c_start).alignment = center
        ws.cell(row=kpi_row_label, column=c_start).fill = kpi_fill
        ws.merge_cells(start_row=kpi_row_label, start_column=c_start, end_row=kpi_row_label, end_column=c_end)

        # 値行
        ws.cell(row=kpi_row_value, column=c_start, value=value).font = kpi_value_font
        ws.cell(row=kpi_row_value, column=c_start).alignment = center
        ws.cell(row=kpi_row_value, column=c_start).fill = kpi_fill
        ws.merge_cells(start_row=kpi_row_value, start_column=c_start, end_row=kpi_row_value, end_column=c_end)

        # 単位 + ヒント
        unit_text = unit + ('　' + hint if hint else '')
        ws.cell(row=kpi_row_unit, column=c_start, value=unit_text).font = kpi_unit_font
        ws.cell(row=kpi_row_unit, column=c_start).alignment = center
        ws.cell(row=kpi_row_unit, column=c_start).fill = kpi_fill
        ws.merge_cells(start_row=kpi_row_unit, start_column=c_start, end_row=kpi_row_unit, end_column=c_end)

        # 枠線（全部）
        for rr in (kpi_row_label, kpi_row_value, kpi_row_unit):
            for cc in range(c_start, c_end + 1):
                ws.cell(row=rr, column=cc).border = kpi_border

    # ----- データ集計（チャート用） -----
    # 1. 日別件数
    date_counts = Counter()
    for r in records:
        d = r['createdAt'][:10] if r['createdAt'] else ''
        if d:
            date_counts[d] += 1
    dates_sorted = sorted(date_counts.keys())

    # 2. 所要時間分布（gen_records のみ）
    bins = [(0, 60, '〜1分'), (60, 300, '1〜5分'), (300, 600, '5〜10分'),
            (600, 1200, '10〜20分'), (1200, 1800, '20〜30分'),
            (1800, 3600, '30〜60分'), (3600, 999999, '60分〜')]
    bin_counts = []
    for low, high, label in bins:
        c = sum(1 for r in gen_records if low <= r['durationSec'] < high)
        bin_counts.append((label, c))

    # 3. ユーザー TOP 10
    user_counts = Counter(r['username'] for r in records)
    top_users = user_counts.most_common(10)

    # 4. スライド数 vs 所要時間（散点）
    scatter_data = [(r['actualSlides'], r['durationSec'] / 60) for r in gen_records if r['actualSlides'] > 0]

    # ----- データ書込み領域（隠しでない、N 列以降に置く） -----
    # Chart 用のソースデータを N 列以降に書く（dashboard sheet 内）
    DATA_COL = 14  # N 列以降
    ws.column_dimensions[get_column_letter(DATA_COL)].width = 12
    for cc in range(DATA_COL + 1, DATA_COL + 7):
        ws.column_dimensions[get_column_letter(cc)].width = 10

    # 各テーブルの開始行
    table_row = 1
    ws.cell(row=table_row, column=DATA_COL, value='日付').font = bold
    ws.cell(row=table_row, column=DATA_COL + 1, value='件数').font = bold
    for i, d in enumerate(dates_sorted):
        ws.cell(row=table_row + 1 + i, column=DATA_COL, value=d)
        ws.cell(row=table_row + 1 + i, column=DATA_COL + 1, value=date_counts[d])
    daily_end = table_row + len(dates_sorted)

    bin_row = daily_end + 3
    ws.cell(row=bin_row, column=DATA_COL, value='時間範囲').font = bold
    ws.cell(row=bin_row, column=DATA_COL + 1, value='件数').font = bold
    for i, (label, cnt) in enumerate(bin_counts):
        ws.cell(row=bin_row + 1 + i, column=DATA_COL, value=label)
        ws.cell(row=bin_row + 1 + i, column=DATA_COL + 1, value=cnt)
    bin_end = bin_row + len(bin_counts)

    user_row = bin_end + 3
    ws.cell(row=user_row, column=DATA_COL, value='ユーザー').font = bold
    ws.cell(row=user_row, column=DATA_COL + 1, value='件数').font = bold
    for i, (uname, cnt) in enumerate(top_users):
        ws.cell(row=user_row + 1 + i, column=DATA_COL, value=uname)
        ws.cell(row=user_row + 1 + i, column=DATA_COL + 1, value=cnt)
    user_end = user_row + len(top_users)

    scatter_row = user_end + 3
    ws.cell(row=scatter_row, column=DATA_COL, value='スライド数').font = bold
    ws.cell(row=scatter_row, column=DATA_COL + 1, value='所要分').font = bold
    for i, (sl, mn) in enumerate(scatter_data):
        ws.cell(row=scatter_row + 1 + i, column=DATA_COL, value=sl)
        ws.cell(row=scatter_row + 1 + i, column=DATA_COL + 1, value=round(mn, 1))
    scatter_end = scatter_row + len(scatter_data)

    # 5. 会社別 件数（当期 / 前期）— 使用意愿度シグナルの元データ
    cur_companies = aggregate(records)['companies']
    prev_companies = prev_agg['companies'] if prev_agg else Counter()
    # 当期 + 前期に登場する全社を当期件数の降順で
    all_companies = sorted(
        set(cur_companies) | set(prev_companies),
        key=lambda c: (-cur_companies.get(c, 0), -prev_companies.get(c, 0))
    )[:12]
    comp_row = scatter_end + 3
    ws.cell(row=comp_row, column=DATA_COL, value='会社').font = bold
    ws.cell(row=comp_row, column=DATA_COL + 1, value='当期').font = bold
    ws.cell(row=comp_row, column=DATA_COL + 2, value='前期').font = bold
    for i, comp in enumerate(all_companies):
        ws.cell(row=comp_row + 1 + i, column=DATA_COL, value=comp)
        ws.cell(row=comp_row + 1 + i, column=DATA_COL + 1, value=cur_companies.get(comp, 0))
        ws.cell(row=comp_row + 1 + i, column=DATA_COL + 2, value=prev_companies.get(comp, 0))
    comp_end = comp_row + len(all_companies)

    # ----- セクション見出し -----
    section_row = 8
    ws.cell(row=section_row, column=1, value='■ チャート分析').font = section_font
    ws.row_dimensions[section_row].height = 22

    # ----- グラフ作成 -----
    # Chart 1: 日別件数（柱状）
    ch1 = BarChart()
    ch1.type = 'col'
    ch1.style = 11
    ch1.title = '日別 PPT 生成件数'
    ch1.x_axis.title = '日付'
    ch1.y_axis.title = '件数'
    ch1.height = 8
    ch1.width = 15
    data_ref = Reference(ws, min_col=DATA_COL + 1, min_row=table_row, max_row=daily_end)
    cats_ref = Reference(ws, min_col=DATA_COL, min_row=table_row + 1, max_row=daily_end)
    ch1.add_data(data_ref, titles_from_data=True)
    ch1.set_categories(cats_ref)
    ch1.legend = None
    ch1.dataLabels = DataLabelList(showVal=True, showCatName=False, showSerName=False, showLegendKey=False, showPercent=False, showBubbleSize=False)
    ws.add_chart(ch1, 'A10')

    # Chart 2: 所要時間分布
    ch2 = BarChart()
    ch2.type = 'col'
    ch2.style = 12
    ch2.title = '所要時間分布（生成成功 ' + str(len(gen_records)) + ' 件）'
    ch2.x_axis.title = '所要時間'
    ch2.y_axis.title = '件数'
    ch2.height = 8
    ch2.width = 15
    data_ref = Reference(ws, min_col=DATA_COL + 1, min_row=bin_row, max_row=bin_end)
    cats_ref = Reference(ws, min_col=DATA_COL, min_row=bin_row + 1, max_row=bin_end)
    ch2.add_data(data_ref, titles_from_data=True)
    ch2.set_categories(cats_ref)
    ch2.legend = None
    ch2.dataLabels = DataLabelList(showVal=True, showCatName=False, showSerName=False, showLegendKey=False, showPercent=False, showBubbleSize=False)
    ws.add_chart(ch2, 'G10')

    # Chart 3: ユーザー TOP 10（横棒）
    ch3 = BarChart()
    ch3.type = 'bar'
    ch3.style = 13
    ch3.title = f'ユーザー別件数 TOP {len(top_users)}'
    ch3.x_axis.title = '件数'
    ch3.y_axis.title = 'ユーザー'
    ch3.height = 9
    ch3.width = 15
    data_ref = Reference(ws, min_col=DATA_COL + 1, min_row=user_row, max_row=user_end)
    cats_ref = Reference(ws, min_col=DATA_COL, min_row=user_row + 1, max_row=user_end)
    ch3.add_data(data_ref, titles_from_data=True)
    ch3.set_categories(cats_ref)
    ch3.legend = None
    ch3.dataLabels = DataLabelList(showVal=True, showCatName=False, showSerName=False, showLegendKey=False, showPercent=False, showBubbleSize=False)
    ws.add_chart(ch3, 'A26')

    # Chart 4: スライド vs 所要時間（散点 — marker のみ、線なし）
    ch4 = ScatterChart()
    ch4.scatterStyle = "marker"  # ★ chart レベルで marker のみ
    ch4.title = 'スライド数 × 所要時間（生成成功）'
    ch4.x_axis.title = 'スライド数'
    ch4.y_axis.title = '所要時間（分）'
    ch4.style = 9
    ch4.height = 9
    ch4.width = 15
    if scatter_data:
        x_ref = Reference(ws, min_col=DATA_COL, min_row=scatter_row + 1, max_row=scatter_end)
        y_ref = Reference(ws, min_col=DATA_COL + 1, min_row=scatter_row + 1, max_row=scatter_end)
        series = Series(y_ref, x_ref, title='所要時間')
        # 線を完全に消す
        gp_series = GraphicalProperties()
        gp_series.line = LineProperties(noFill=True)
        series.graphicalProperties = gp_series
        # マーカー（点）の見た目を明示
        marker = Marker(symbol='circle', size=5)
        gp_marker = GraphicalProperties(solidFill='1F4E78')
        gp_marker.line = LineProperties(solidFill='1F4E78')
        marker.spPr = gp_marker
        series.marker = marker
        ch4.series.append(series)
    ch4.legend = None
    ws.add_chart(ch4, 'G26')

    # Chart 5: 会社別件数（当期 vs 前期）— 使用意愿度トレンド
    ch5 = BarChart()
    ch5.type = 'bar'
    ch5.style = 10
    ch5.title = '会社別件数（当期 vs 前期）' if prev_agg else '会社別件数（当期）'
    ch5.x_axis.title = '件数'
    ch5.y_axis.title = '会社'
    ch5.height = 9
    ch5.width = 15
    max_data_col = DATA_COL + 2 if prev_agg else DATA_COL + 1
    data_ref = Reference(ws, min_col=DATA_COL + 1, max_col=max_data_col, min_row=comp_row, max_row=comp_end)
    cats_ref = Reference(ws, min_col=DATA_COL, min_row=comp_row + 1, max_row=comp_end)
    ch5.add_data(data_ref, titles_from_data=True)
    ch5.set_categories(cats_ref)
    ch5.dataLabels = DataLabelList(showVal=True, showCatName=False, showSerName=False, showLegendKey=False, showPercent=False, showBubbleSize=False)
    if not prev_agg:
        ch5.legend = None
    ws.add_chart(ch5, 'A42')

    # ----- 使用意愿度シグナル（会社別 前期比）-----
    sig_row = 42
    if prev_agg:
        ws.cell(row=sig_row, column=8, value='■ 使用意愿度シグナル（会社別 前期比）').font = section_font
        ws.merge_cells(start_row=sig_row, start_column=8, end_row=sig_row, end_column=12)
        ws.row_dimensions[sig_row].height = 22
        sig_hdr = sig_row + 1
        for ci, htxt in enumerate(['会社', '当期', '前期', '増減', 'シグナル']):
            cell = ws.cell(row=sig_hdr, column=8 + ci, value=htxt)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', start_color='305496')
            cell.alignment = center
        # 増減の大きい順（活性化・離脱の両端）に並べる
        signal_rows = []
        for comp in all_companies:
            cur = cur_companies.get(comp, 0)
            prev = prev_companies.get(comp, 0)
            diff = cur - prev
            if diff > 0:
                sig = '▲ 利用拡大'
            elif diff < 0:
                sig = '▼ 利用縮小（要フォロー）' if prev >= 3 else '▼ 微減'
            else:
                sig = '± 横ばい'
            signal_rows.append((comp, cur, prev, diff, sig))
        signal_rows.sort(key=lambda x: -abs(x[3]))
        green = Font(color='2E7D32', bold=True)
        red = Font(color='C00000', bold=True)
        for i, (comp, cur, prev, diff, sig) in enumerate(signal_rows[:8]):
            rr = sig_hdr + 1 + i
            ws.cell(row=rr, column=8, value=comp)
            ws.cell(row=rr, column=9, value=cur).alignment = center
            ws.cell(row=rr, column=10, value=prev).alignment = center
            dc = ws.cell(row=rr, column=11, value=(f'+{diff}' if diff > 0 else str(diff)))
            dc.alignment = center
            dc.font = green if diff > 0 else (red if diff < 0 else bold)
            sc = ws.cell(row=rr, column=12, value=sig)
            sc.font = green if diff > 0 else (red if diff < 0 else Font(color='808080'))

    # ----- 完成度ブロック（下）-----
    stat_row = 54
    ws.cell(row=stat_row, column=1, value='■ 完成度サマリ').font = section_font
    ws.row_dimensions[stat_row].height = 22

    stat_table = [
        ('総体', total_rows, '本', ''),
        ('統計上の未生成', excluded_count, '本', '（status="created" or 空 かつ logsCount≤1）'),
        ('調査後実際の未生成', excluded_count, '本', '（手動で更新）'),
    ]
    if is_v2:
        stat_table.append(('中断のあった件数', break_count_total, '本', '（log 間隔 ≥ 1 時間、薄黄ハイライト）'))
    stat_table.extend([
        ('調査前完成率', f'{completion_rate*100:.1f}%', '', ''),
        ('調査後完成率', f'{completion_rate*100:.1f}%', '', '（調査後未生成を更新したら再計算）'),
    ])

    header_stat_font = Font(bold=True, color='FFFFFF')
    header_stat_fill = PatternFill('solid', start_color='305496')
    table_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    label_fill = PatternFill('solid', start_color=COLOR_KPI_BG)

    ws.cell(row=stat_row + 2, column=1, value='指標').font = header_stat_font
    ws.cell(row=stat_row + 2, column=2, value='値').font = header_stat_font
    ws.cell(row=stat_row + 2, column=3, value='単位').font = header_stat_font
    ws.cell(row=stat_row + 2, column=4, value='備考').font = header_stat_font
    for cc in range(1, 5):
        ws.cell(row=stat_row + 2, column=cc).fill = header_stat_fill
        ws.cell(row=stat_row + 2, column=cc).alignment = center
        ws.cell(row=stat_row + 2, column=cc).border = table_border
    ws.merge_cells(start_row=stat_row + 2, start_column=4, end_row=stat_row + 2, end_column=8)

    for i, (label, value, unit, note) in enumerate(stat_table):
        rr = stat_row + 3 + i
        ws.cell(row=rr, column=1, value=label).font = bold
        ws.cell(row=rr, column=1).fill = label_fill
        ws.cell(row=rr, column=2, value=value).alignment = center
        ws.cell(row=rr, column=3, value=unit).alignment = center
        ws.cell(row=rr, column=4, value=note).font = Font(italic=True, color='808080')
        for cc in range(1, 5):
            ws.cell(row=rr, column=cc).border = table_border
        ws.merge_cells(start_row=rr, start_column=4, end_row=rr, end_column=8)

    # ----- 時間定義の注記 -----
    if is_v2:
        note_row = stat_row + 4 + len(stat_table)
        ws.cell(row=note_row, column=1,
                value='※ 「純所要時間」= log 間隔 ≥ 1 時間の中断を除外した実生成時間。中断ありの行は薄黄でハイライト。').font = Font(italic=True, color='808080', size=10)
        ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=12)


def build(input_tsv: str, output_xlsx: str, start_date: str = '', end_date: str = '',
          prev_tsv: str = '', mode: str = 'weekly'):
    records, is_v2 = parse_tsv(input_tsv)

    prev_agg = None
    if prev_tsv:
        try:
            prev_records, _ = parse_tsv(prev_tsv)
            prev_agg = aggregate(prev_records)
        except Exception as e:  # 前期データが欠けても本体は出す
            print(f'  [warn] prev tsv の読込に失敗、前期比をスキップ: {e}')

    wb = Workbook()
    ws_dash = wb.active
    ws_dash.title = 'ダッシュボード'

    ws_detail = wb.create_sheet('明細データ')
    detail_stats = build_detail_sheet(ws_detail, records, is_v2)

    build_dashboard(ws_dash, records, is_v2, start_date, end_date, detail_stats,
                    prev_agg=prev_agg, mode=mode)

    wb.save(output_xlsx)
    print(f'Saved: {output_xlsx}')
    print(f'  mode: {mode} / {"v2 (net time)" if is_v2 else "v1 (raw time)"}')
    print(f'  total rows: {detail_stats["total_rows"]}')
    print(f'  excluded (J=0): {detail_stats["excluded_count"]}')
    print(f'  generated (J>0): {detail_stats["total_rows"] - detail_stats["excluded_count"]}')
    if is_v2:
        print(f'  rows with breaks (>1h gap): {detail_stats["break_count_total"]}')
    if prev_agg:
        print(f'  prev period: {prev_agg["total"]} 件 / 完成率 {prev_agg["completion"]*100:.1f}%（前期比を Dashboard に反映）')


def _parse_args(argv):
    """位置引数 + --prev/--mode のフラグをパース。"""
    pos, prev_tsv, mode = [], '', 'weekly'
    i = 0
    while i < len(argv):
        a = argv[i]
        if a == '--prev':
            prev_tsv = argv[i + 1]; i += 2; continue
        if a == '--mode':
            mode = argv[i + 1]; i += 2; continue
        pos.append(a); i += 1
    while len(pos) < 4:
        pos.append('')
    return pos[0], pos[1], pos[2], pos[3], prev_tsv, mode


if __name__ == '__main__':
    if len(sys.argv) < 3:
        print('Usage: python3 generate_xlsx.py <input.tsv> <output.xlsx> '
              '[start_date] [end_date] [--prev <prev.tsv>] [--mode daily|weekly]')
        sys.exit(1)
    inp, out, sd, ed, prev, md = _parse_args(sys.argv[1:])
    build(inp, out, sd, ed, prev, md)
